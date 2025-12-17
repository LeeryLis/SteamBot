import json
from pathlib import Path
from collections import OrderedDict
from typing import Dict
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import TextAxis
from openpyxl.chart.marker import Marker
from openpyxl.utils import get_column_letter


class SummarizeToExcel:
    COLUMN_LABELS = {
        "item_name": "Предмет",
        "total_bought": "Куплено (количество)",
        "total_sold": "Продано (количество)",
        "sum_bought": "Потрачено",
        "sum_sold": "Получено",
        "quantity_difference": "Разница (шт.)",
        "sum_difference": "Разница (денег)",
        "sold_to_bought_pct": "Получено/Потрачено (%)",
    }

    MONTHLY_COLUMN_LABELS = {
        "month_str": "Месяц",
        "total_bought": "Куплено (шт.)",
        "total_sold": "Продано (шт.)",
        "sum_bought": "Потрачено",
        "sum_sold": "Получено",
        "quantity_difference": "Разница (шт.)",
        "sum_difference": "Разница (денег)",
    }

    @staticmethod
    def _safe_sheet_name(name: str, existing: set, max_len: int = 31) -> str:
        name = name.replace(":", "")
        base = name[:max_len].strip()
        if base not in existing:
            existing.add(base)
            return base
        idx = 1
        while True:
            suffix = f"_{idx}"
            allowed_base_len = max_len - len(suffix)
            candidate = (name[:allowed_base_len] + suffix).strip()
            if candidate not in existing:
                existing.add(candidate)
                return candidate
            idx += 1

    @staticmethod
    def _load_aggregated(json_path: Path) -> (Dict[str, Dict], Dict[str, str]):
        with open(json_path, "r", encoding="utf-8") as f:
            saved = json.load(f)
        return saved.get("aggregated_data", {}), saved.get("app_id_to_game_name", {})

    @staticmethod
    def _prepare_rows_from_items(items: Dict) -> pd.DataFrame:
        rows = []
        for item_hash_name, stats in items.items():
            sum_bought = round(float(stats.get("sum_bought", 0.0)), 2)
            sum_sold = round(float(stats.get("sum_sold", 0.0)), 2)

            if sum_bought > 0 and sum_sold >= 0:
                ratio = sum_sold / sum_bought
            else:
                ratio = None

            rows.append({
                "item_name": stats.get("item_name"),
                "total_bought": int(stats.get("total_bought", 0)),
                "total_sold": int(stats.get("total_sold", 0)),
                "sum_bought": sum_bought,
                "sum_sold": sum_sold,
                "quantity_difference": int(stats.get("quantity_difference", 0)),
                "sum_difference": round(float(stats.get("sum_difference", 0.0)), 2),
                "sold_to_bought_pct": ratio,
            })

        cols = ["item_name", "total_bought", "total_sold", "sum_bought", "sum_sold",
                "quantity_difference", "sum_difference", "sold_to_bought_pct"]

        if not rows:
            return pd.DataFrame(columns=cols)
        df = pd.DataFrame(rows)[cols]
        df = df.sort_values("item_name").reset_index(drop=True)
        return df

    @staticmethod
    def _calc_column_width(series: pd.Series, display_name: str | None = None) -> int:
        header = display_name if display_name is not None else str(series.name)
        max_val_len = series.astype(str).map(len).max() if len(series) > 0 else 0
        max_len = max(max_val_len, len(header))
        return min(max(14, int(max_len * 1.4) + 4), 90)

    def _apply_number_formats(self, worksheet: Worksheet, df: pd.DataFrame, column_labels: dict[str, str]) -> None:
        money_cols = {"sum_bought", "sum_sold", "sum_difference"}
        int_cols = {"total_bought", "total_sold", "quantity_difference"}

        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)

            display_name = column_labels.get(col_name, col_name)

            width = self._calc_column_width(df[col_name], display_name=display_name)
            worksheet.column_dimensions[col_letter].width = width

            if col_name in money_cols:
                fmt = '#,##0.00'
            elif col_name in int_cols:
                fmt = '0'
            elif col_name == "sold_to_bought_pct":
                fmt = '0.00%'
            else:
                fmt = None

            if fmt is not None:
                for row_idx in range(2, 2 + len(df)):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        if col_name == "sold_to_bought_pct" and (pd.isna(cell.value) or cell.value is None):
                            cell.value = None
                        else:
                            cell.number_format = fmt

    @staticmethod
    def _apply_header_and_freeze(worksheet: Worksheet, df: pd.DataFrame) -> None:
        last_row = len(df) + 1
        last_col_letter = get_column_letter(len(df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
        worksheet.freeze_panes = "A2"

        header_font = Font(bold=True)
        for col_idx in range(1, len(df.columns) + 1):
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.alignment = Alignment(vertical="center", horizontal="center")

    @staticmethod
    def _apply_conditional_formatting(worksheet: Worksheet, df: pd.DataFrame) -> None:
        last_row = len(df) + 1

        red_fill = PatternFill(start_color="FFF4C7C7", end_color="FFF4C7C7", fill_type="solid")
        green_fill = PatternFill(start_color="FFDCF7DC", end_color="FFDCF7DC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")

        try:
            sd_col_idx = list(df.columns).index("sum_difference") + 1
            sd_col_letter = get_column_letter(sd_col_idx)
            sd_range = f"{sd_col_letter}2:{sd_col_letter}{last_row}"
            worksheet.conditional_formatting.add(
                sd_range,
                CellIsRule(operator='lessThanOrEqual', formula=['0'], stopIfTrue=True, fill=red_fill)
            )
            worksheet.conditional_formatting.add(
                sd_range,
                CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill)
            )
        except ValueError:
            pass

        try:
            pct_col_idx = list(df.columns).index("sold_to_bought_pct") + 1
            pct_col_letter = get_column_letter(pct_col_idx)
            pct_range = f"{pct_col_letter}2:{pct_col_letter}{last_row}"

            # <=100%
            worksheet.conditional_formatting.add(
                pct_range,
                FormulaRule(formula=[f'=AND(NOT(ISBLANK({pct_col_letter}2)),{pct_col_letter}2<=1)'],
                            fill=red_fill, stopIfTrue=True)
            )
            # >100% and <=104%
            worksheet.conditional_formatting.add(
                pct_range,
                FormulaRule(
                    formula=[f'=AND(NOT(ISBLANK({pct_col_letter}2)),{pct_col_letter}2>1,{pct_col_letter}2<=1.04)'],
                    fill=yellow_fill, stopIfTrue=True)
            )
            # >104%
            worksheet.conditional_formatting.add(
                pct_range,
                FormulaRule(formula=[f'=AND(NOT(ISBLANK({pct_col_letter}2)),{pct_col_letter}2>1.04)'],
                            fill=green_fill)
            )
        except ValueError:
            pass

    @staticmethod
    def _write_summary(worksheet, df: pd.DataFrame, start_col: int) -> None:
        total_sum_bought = round(df["sum_bought"].sum() if "sum_bought" in df.columns else 0.0, 2)
        total_sum_sold = round(df["sum_sold"].sum() if "sum_sold" in df.columns else 0.0, 2)
        total_sum_diff = round(df["sum_difference"].sum() if "sum_difference" in df.columns else 0.0, 2)

        if total_sum_bought > 0 and total_sum_sold >= 0:
            total_pct = total_sum_sold / total_sum_bought
        else:
            total_pct = None

        summary_rows = [
            ("Потрачено", total_sum_bought, "money"),
            ("Получено", total_sum_sold, "money"),
            ("Разница", total_sum_diff, "money"),
            ("Получено/Потрачено (%)", total_pct, "pct"),
        ]

        worksheet.cell(row=1, column=start_col, value="Итог").font = Font(bold=True)
        try:
            worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+1)
        except Exception:
            pass

        for i, (label, value, value_type) in enumerate(summary_rows, start=1):
            worksheet.cell(row=i+1, column=start_col, value=label)
            cell_val = worksheet.cell(row=i+1, column=start_col + 1, value=value)
            if value is None:
                cell_val.value = None
            else:
                if value_type == "money":
                    cell_val.number_format = "#,##0.00"
                elif value_type == "pct":
                    cell_val.number_format = '0.00%'
            cell_val.alignment = Alignment(horizontal="right")

        worksheet.column_dimensions[get_column_letter(start_col)].width = 28
        worksheet.column_dimensions[get_column_letter(start_col + 1)].width = 18

        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        summary_max_row = 1 + len(summary_rows)
        summary_max_col = start_col + 1
        for r in range(1, summary_max_row + 1):
            for c in range(start_col, summary_max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                left = medium if c == start_col else thin
                right = medium if c == summary_max_col else thin
                top = medium if r == 1 else thin
                bottom = medium if r == summary_max_row else thin
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    @staticmethod
    def _apply_table_borders(worksheet: Worksheet, df: pd.DataFrame) -> None:
        thin = Side(border_style="thin", color="000000")
        medium = Side(border_style="medium", color="000000")
        thick = Side(border_style="thick", color="000000")

        max_row = len(df) + 1
        max_col = len(df.columns)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                left = thick if c == 1 else medium
                right = thick if c == max_col else medium
                top = thick if r == 1 else thin
                bottom = thick if r == max_row else thin
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def summarize_json_to_excel(self, json_path_str: str, excel_path_str: str) -> None:
        json_path = Path(json_path_str)
        excel_path = Path(excel_path_str)
        if not json_path.exists():
            raise FileNotFoundError(f"Файл не найден: {json_path}")

        aggregated, app_id_to_game_name = self._load_aggregated(json_path)
        if not aggregated:
            raise ValueError("В файле нет aggregated_data или он пуст.")

        games = OrderedDict(sorted(aggregated.items(), key=lambda kv: kv[0].lower()))
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(excel_path, engine="openpyxl", date_format="YYYY-MM-DD") as writer:
            existing_sheet_names = set()

            for app_id, items in games.items():
                game_name = app_id_to_game_name.get(app_id)
                df = self._prepare_rows_from_items(items)
                df_export = df.rename(columns=self.COLUMN_LABELS)

                sheet_name = self._safe_sheet_name(game_name, existing_sheet_names)
                df_export.to_excel(writer, sheet_name=sheet_name, index=False)
                worksheet = writer.sheets[sheet_name]

                self._apply_number_formats(worksheet, df, self.COLUMN_LABELS)
                self._apply_header_and_freeze(worksheet, df)
                self._apply_conditional_formatting(worksheet, df)

                summary_start_col = len(df.columns) + 3
                self._write_summary(worksheet, df, summary_start_col)

                self._apply_table_borders(worksheet, df)

        print(f"Excel сохранён: {excel_path_str}")

    @staticmethod
    def _parse_month_key(key: str) -> datetime:
        return datetime.strptime(key, "%m.%Y")

    @staticmethod
    def _prepare_rows_from_months(months: Dict) -> pd.DataFrame:
        rows = []

        for month_key, stats in months.items():
            month_dt = datetime.strptime(month_key, "%m.%Y")

            rows.append({
                "month": month_dt,
                "month_str": month_key,
                "total_bought": int(stats.get("total_bought", 0)),
                "total_sold": int(stats.get("total_sold", 0)),
                "sum_bought": round(float(stats.get("sum_bought", 0.0)), 2),
                "sum_sold": round(float(stats.get("sum_sold", 0.0)), 2),
                "quantity_difference": int(stats.get("quantity_difference", 0)),
                "sum_difference": round(float(stats.get("sum_difference", 0.0)), 2),
            })

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows)
        df = df.sort_values("month").reset_index(drop=True)

        cols = [
            "month_str",
            "total_bought",
            "total_sold",
            "sum_bought",
            "sum_sold",
            "quantity_difference",
            "sum_difference",
        ]

        return df[cols]

    @staticmethod
    def _add_charts(worksheet, df):
        if df.empty:
            return

        max_row = len(df) + 1

        cats1 = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=max_row
        )

        chart1 = LineChart()
        chart1.title = "Динамика по месяцам"

        chart1.height = 15
        chart1.width = 30
        chart1.legend.position = "r"
        chart1.legend.overlay = False

        chart1.x_axis = TextAxis()
        chart1.x_axis.tickLblPos = "nextTo"
        chart1.x_axis.majorTickMark = "out"
        chart1.x_axis.delete = False

        chart1.y_axis.delete = False

        for col_name in ("sum_bought", "sum_sold"):
            col_idx = list(df.columns).index(col_name) + 1
            data1 = Reference(
                worksheet,
                min_col=col_idx,
                min_row=1,
                max_row=max_row
            )
            chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)

        for s in chart1.series:
            s.marker = Marker(symbol="circle", size=5)
            s.smooth = False

        cats2 = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=max_row
        )

        chart2 = LineChart()
        chart2.title = "Разница по месяцам"
        chart2.height = 15
        chart2.width = 30
        chart2.legend = None

        chart2.x_axis = TextAxis()
        chart2.x_axis.tickLblPos = "nextTo"
        chart2.x_axis.majorTickMark = "out"
        chart2.x_axis.delete = False

        chart2.y_axis.delete = False

        col_idx = list(df.columns).index("sum_difference") + 1
        data2 = Reference(
            worksheet,
            min_col=col_idx,
            min_row=2,
            max_row=max_row
        )
        chart2.add_data(data2, titles_from_data=False)
        chart2.set_categories(cats2)

        chart2.series[0].marker = Marker(symbol="circle", size=5)
        chart2.series[0].smooth = False

        data_fake = Reference(worksheet, min_col=col_idx, min_row=2, max_row=2)
        chart2.add_data(data_fake, titles_from_data=False)

        chart_col = len(df.columns) + 3

        chart1_anchor = f"{get_column_letter(chart_col)}2"
        worksheet.add_chart(chart1, chart1_anchor)

        chart2_anchor = f"{get_column_letter(chart_col)}{2 + 30}"
        worksheet.add_chart(chart2, chart2_anchor)

    @staticmethod
    def _apply_monthly_conditional_formatting(worksheet: Worksheet, df: pd.DataFrame) -> None:
        last_row = len(df) + 1

        green_fill = PatternFill(start_color="FFDCF7DC", end_color="FFDCF7DC", fill_type="solid")
        red_fill = PatternFill(start_color="FFF4C7C7", end_color="FFF4C7C7", fill_type="solid")

        try:
            col_idx = list(df.columns).index("sum_difference") + 1
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{last_row}"

            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill)
            )
            worksheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
            )
        except ValueError:
            pass

    def monthly_summarize_json_to_excel(self, json_path_str: str, excel_path_str: str) -> None:
        json_path = Path(json_path_str)
        excel_path = Path(excel_path_str)

        aggregated, app_id_to_game_name = self._load_aggregated(json_path)

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            used_names = set()

            for app_id, months in aggregated.items():
                game_name = app_id_to_game_name.get(app_id, app_id)

                df = self._prepare_rows_from_months(months)
                df_export = df.rename(columns=self.MONTHLY_COLUMN_LABELS)

                sheet = self._safe_sheet_name(game_name, used_names)
                df_export.to_excel(writer, sheet_name=sheet, index=False)
                ws = writer.sheets[sheet]

                self._apply_number_formats(ws, df, self.MONTHLY_COLUMN_LABELS)
                self._apply_header_and_freeze(ws, df)
                self._apply_monthly_conditional_formatting(ws, df)
                self._apply_table_borders(ws, df)
                self._add_charts(ws, df)

        print(f"Excel сохранён: {excel_path_str}")
